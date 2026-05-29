from PyQt6 import QtWidgets, QtCore, QtGui
import sys
from bycle_db import seed_default, search_items, insert_item, update_item, delete_item, init_db
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
from datetime import datetime


class BycleManagerApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        init_db()
        seed_default(100)
        self.setWindowTitle('Bycle Manager')
        self.resize(900, 600)
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)

        form = QtWidgets.QHBoxLayout()
        self.id_label = QtWidgets.QLabel('ID:')
        self.id_display = QtWidgets.QLabel('')
        self.name_input = QtWidgets.QLineEdit()
        self.name_input.setPlaceholderText('자전거 이름')
        self.price_input = QtWidgets.QLineEdit()
        self.price_input.setPlaceholderText('가격')
        self.qty_input = QtWidgets.QLineEdit()
        self.qty_input.setPlaceholderText('수량')

        form.addWidget(self.id_label)
        form.addWidget(self.id_display)
        form.addWidget(QtWidgets.QLabel('이름:'))
        form.addWidget(self.name_input)
        form.addWidget(QtWidgets.QLabel('가격:'))
        form.addWidget(self.price_input)
        form.addWidget(QtWidgets.QLabel('수량:'))
        form.addWidget(self.qty_input)

        btns = QtWidgets.QHBoxLayout()
        self.add_btn = QtWidgets.QPushButton('추가')
        self.update_btn = QtWidgets.QPushButton('수정')
        self.delete_btn = QtWidgets.QPushButton('삭제')
        self.clear_btn = QtWidgets.QPushButton('초기화')
        self.excel_btn = QtWidgets.QPushButton('엑셀로 출력')

        btns.addWidget(self.add_btn)
        btns.addWidget(self.update_btn)
        btns.addWidget(self.delete_btn)
        btns.addWidget(self.clear_btn)
        btns.addWidget(self.excel_btn)

        search_layout = QtWidgets.QHBoxLayout()
        self.search_input = QtWidgets.QLineEdit()
        self.search_input.setPlaceholderText('검색어 입력 (이름)')
        self.search_btn = QtWidgets.QPushButton('검색')
        self.refresh_btn = QtWidgets.QPushButton('전체')
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        search_layout.addWidget(self.refresh_btn)

        layout.addLayout(form)
        layout.addLayout(btns)
        layout.addLayout(search_layout)

        # Table
        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['ID', 'Name', 'Price', 'Qty'])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        # connections
        self.add_btn.clicked.connect(self.add_item)
        self.update_btn.clicked.connect(self.update_item)
        self.delete_btn.clicked.connect(self.delete_item)
        self.clear_btn.clicked.connect(self.clear_inputs)
        self.search_btn.clicked.connect(self.search)
        self.refresh_btn.clicked.connect(self.load_data)
        self.excel_btn.clicked.connect(self.export_to_excel)
        self.table.cellDoubleClicked.connect(self.load_selected_to_form)

        # styling
        self.apply_styles()

    def apply_styles(self):
        qss = """
        QWidget { background: #0f1724; color: #e6eef8; font-family: 'Segoe UI', Arial; }
        QLabel { font-weight: 600; }
        QLineEdit { background: #0b1220; border: 1px solid #243b55; padding:6px; border-radius:4px; }
        QPushButton { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2b94f6, stop:1 #1b6fd6); color: white; padding:8px 12px; border-radius:6px; }
        QPushButton:hover { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #3aa6ff, stop:1 #1e7fe0); }
        QTableWidget { background: #0b1220; gridline-color: #223547; }
        QHeaderView::section { background: #17293b; color: #cfe9ff; padding:6px; }
        QTableWidget::item:selected { background: #2b6f9e; color: #fff; }
        """
        self.setStyleSheet(qss)

    def load_data(self):
        rows = search_items('')
        self.populate_table(rows)

    def populate_table(self, rows):
        self.table.setRowCount(0)
        for r in rows:
            row_idx = self.table.rowCount()
            self.table.insertRow(row_idx)
            for c, val in enumerate(r):
                item = QtWidgets.QTableWidgetItem(str(val))
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_idx, c, item)

    def load_selected_to_form(self, row, col):
        item_id = int(self.table.item(row, 0).text())
        name = self.table.item(row, 1).text()
        price = self.table.item(row, 2).text()
        qty = self.table.item(row, 3).text()
        self.id_display.setText(str(item_id))
        self.name_input.setText(name)
        self.price_input.setText(price)
        self.qty_input.setText(qty)

    def add_item(self):
        name = self.name_input.text().strip()
        try:
            price = int(self.price_input.text())
            qty = int(self.qty_input.text())
        except ValueError:
            QtWidgets.QMessageBox.warning(self, '입력 오류', '가격과 수량은 숫자여야 합니다.')
            return
        if not name:
            QtWidgets.QMessageBox.warning(self, '입력 오류', '이름을 입력하세요.')
            return
        insert_item(name, price, qty)
        self.load_data()
        self.clear_inputs()

    def update_item(self):
        try:
            item_id = int(self.id_display.text())
        except ValueError:
            QtWidgets.QMessageBox.warning(self, '선택 오류', '수정할 항목을 선택하세요.')
            return
        name = self.name_input.text().strip()
        try:
            price = int(self.price_input.text())
            qty = int(self.qty_input.text())
        except ValueError:
            QtWidgets.QMessageBox.warning(self, '입력 오류', '가격과 수량은 숫자여야 합니다.')
            return
        update_item(item_id, name, price, qty)
        self.load_data()
        self.clear_inputs()

    def delete_item(self):
        try:
            item_id = int(self.id_display.text())
        except ValueError:
            QtWidgets.QMessageBox.warning(self, '선택 오류', '삭제할 항목을 선택하세요.')
            return
        reply = QtWidgets.QMessageBox.question(self, '확인', '정말 삭제하시겠습니까?', QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No)
        if reply == QtWidgets.QMessageBox.StandardButton.Yes:
            delete_item(item_id)
            self.load_data()
            self.clear_inputs()

    def clear_inputs(self):
        self.id_display.setText('')
        self.name_input.clear()
        self.price_input.clear()
        self.qty_input.clear()

    def search(self):
        kw = self.search_input.text().strip()
        rows = search_items(kw)
        self.populate_table(rows)

    def export_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Bycle'

            # 헤더
            headers = ['ID', 'Name', 'Price', 'Qty']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 데이터
            for row_idx in range(self.table.rowCount()):
                for col_idx in range(self.table.columnCount()):
                    cell_text = self.table.item(row_idx, col_idx).text()
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=cell_text)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10

            # 파일명
            filename = f"bycle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.path.dirname(__file__), filename)
            wb.save(filepath)
            
            QtWidgets.QMessageBox.information(self, '성공', f'엑셀 파일이 저장되었습니다.\n{filename}')
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, '오류', f'엑셀 저장 중 오류가 발생했습니다.\n{str(e)}')


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = BycleManagerApp()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
